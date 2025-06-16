import os
import pandas as pd
import customtkinter as ctk
import tkinter.filedialog as fd
import tkinter.messagebox as mb
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

ctk.set_appearance_mode("system")
ctk.set_default_color_theme("blue")

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

BANK_COLS = ["Post Date", "Check", "Description", "Debit", "Credit"]
QB_COLS = ["Type", "Date", "Num", "Name", "Memo", "Clr", "Split", "Amount"]

def reconcile(bank_path: str, qb_path: str, output_path: str) -> None:
    bank = pd.read_csv(bank_path, parse_dates=["Post Date"])[BANK_COLS].copy()

    qb_raw = pd.read_csv(qb_path, parse_dates=["Date"])[QB_COLS].copy()
    qb = qb_raw[
        (qb_raw["Clr"].fillna("").str.strip() == "") &                # only blank Clr
        (qb_raw["Type"].fillna("").astype(str).str.strip() != "") &   # must have Type
        (qb_raw["Date"].notna())                                      # must have Date
    ].copy()

    bank["Credit"] = pd.to_numeric(bank["Credit"], errors="coerce").round(2)
    qb["Amount"] = pd.to_numeric(qb["Amount"], errors="coerce").round(2)

    bank.sort_values("Credit", inplace=True, ignore_index=True)
    qb.sort_values("Amount", inplace=True, ignore_index=True)

    bank_counts = bank["Credit"].value_counts().to_dict()
    qb_counts = qb["Amount"].value_counts().to_dict()

    extras_idx = []
    extras_records = []
    for amt, q_cnt in qb_counts.items():
        b_cnt = bank_counts.get(amt, 0)
        if q_cnt > b_cnt:
            diff = q_cnt - b_cnt
            rows_amt = qb[qb["Amount"] == amt]
            # prioritise removing rows whose Split does NOT look like a file reference YYYY-N
            non_file = rows_amt[~rows_amt["Split"].astype(str).str.match(r"^\d{4}-\d+$", na=False)]
            take = non_file.head(diff)
            if len(take) < diff:
                rem = diff - len(take)
                file_like = rows_amt.loc[rows_amt.index.difference(take.index)].head(rem)
                take = pd.concat([take, file_like])
            extras_idx.extend(take.index.tolist())
            extras_records.extend(take.to_dict("records"))

    qb_clean = qb.drop(extras_idx).copy()

    qb_clean_counts = qb_clean["Amount"].value_counts().to_dict()
    missing_rows = []
    for amt, b_cnt in bank_counts.items():
        q_cnt = qb_clean_counts.get(amt, 0)
        if b_cnt > q_cnt:
            diff = b_cnt - q_cnt
            src = bank[bank["Credit"] == amt]
            # build a counter of how many times each (Amount, Date) already exists in QB
            qb_date_counts = (
                qb_clean[qb_clean["Amount"] == amt]["Date"]
                .dt.date
                .value_counts()
                .to_dict()
            )

            unmatched_rows = []
            for _, r_bank in src.iterrows():
                d = r_bank["Post Date"].date()
                if qb_date_counts.get(d, 0) > 0:
                    qb_date_counts[d] -= 1  # this bank entry is already matched
                else:
                    unmatched_rows.append(r_bank)

            # we only need `diff` rows to add
            unmatched = pd.DataFrame(unmatched_rows).head(diff)
            if len(unmatched) < diff:
                # pad with first rows of src (duplicates allowed) to reach exact diff
                need = diff - len(unmatched)
                unmatched = pd.concat([unmatched, src.head(need)])
            dup = diff > 1
            for _, r in unmatched.iterrows():
                date_val = r["Post Date"].date()
                # collect all unique descriptions for this amount on this date
                descs = (
                    src[src["Post Date"].dt.date == date_val]["Description"]
                    .dropna()
                    .astype(str)
                    .unique()
                )
                split_val = "; ".join(descs)
                missing_rows.append({
                    "Type": "Bank-Add",
                    "Date": r["Post Date"],
                    "Num": "",
                    "Name": "",
                    "Memo": "Duplicate, investigate added date" if dup else "Added to match bank",
                    "Clr": "",
                    "Split": split_val,
                    "Amount": r["Credit"],
                })

    qb_reconciled = pd.concat([qb_clean, pd.DataFrame(missing_rows)], ignore_index=True)
    qb_reconciled.sort_values("Amount", inplace=True, ignore_index=True)

    extras_df = pd.DataFrame(extras_records)
    if not extras_df.empty:
        extras_df.sort_values("Amount", inplace=True, ignore_index=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    spacer = 2
    bank_col0 = 1
    qb_col0 = bank_col0 + len(BANK_COLS) + spacer
    ext_col0 = qb_col0 + len(QB_COLS) + spacer

    for i, col in enumerate(BANK_COLS, bank_col0):
        ws.cell(row=1, column=i, value=col)
    for i, col in enumerate(QB_COLS, qb_col0):
        ws.cell(row=1, column=i, value=col)
    if not extras_df.empty:
        ws.cell(row=1, column=ext_col0 - 1, value="Extras from QB")
        for i, col in enumerate(QB_COLS, ext_col0):
            ws.cell(row=1, column=i, value=col)

    for r_idx, row in enumerate(dataframe_to_rows(bank, index=False, header=False), 2):
        for c_idx, val in enumerate(row, bank_col0):
            ws.cell(row=r_idx, column=c_idx, value=val)

    for r_idx, row in enumerate(dataframe_to_rows(qb_reconciled, index=False, header=False), 2):
        for c_idx, val in enumerate(row, qb_col0):
            ws.cell(row=r_idx, column=c_idx, value=val)

    if not extras_df.empty:
        for r_idx, row in enumerate(dataframe_to_rows(extras_df, index=False, header=False), 2):
            for c_idx, val in enumerate(row, ext_col0):
                ws.cell(row=r_idx, column=c_idx, value=val)

    # highlight spacer columns F and G for visual separation
    spacer_cols = [bank_col0 + len(BANK_COLS), bank_col0 + len(BANK_COLS) + 1]  # 6 and 7
    for r in range(1, ws.max_row + 1):
        for c in spacer_cols:
            ws.cell(row=r, column=c).fill = YELLOW

    # --- Totals rows ---
    bank_total_row = len(bank) + 2
    ws.cell(row=bank_total_row, column=bank_col0 + BANK_COLS.index("Description"), value="TOTAL")
    ws.cell(row=bank_total_row, column=bank_col0 + BANK_COLS.index("Credit"), value=bank["Credit"].sum())

    qb_total_row = len(qb_reconciled) + 2
    ws.cell(row=qb_total_row, column=qb_col0 + QB_COLS.index("Memo"), value="TOTAL")
    ws.cell(row=qb_total_row, column=qb_col0 + QB_COLS.index("Amount"), value=qb_reconciled["Amount"].sum())

    dup_counts = qb_reconciled[qb_reconciled["Type"] == "Bank-Add"]["Amount"].value_counts().to_dict()
    for idx, row in qb_reconciled.iterrows():
        if row["Type"] == "Bank-Add":
            excel_row = idx + 2
            multi_split = isinstance(row["Split"], str) and ";" in row["Split"]
            fill = YELLOW if (dup_counts.get(row["Amount"], 0) > 1 or multi_split) else GREEN
            for col in range(qb_col0, qb_col0 + len(QB_COLS)):
                ws.cell(row=excel_row, column=col).fill = fill
            mask = (bank["Credit"] == row["Amount"]) & (bank["Post Date"].dt.date == row["Date"].date())
            for b_idx in bank[mask].index:
                for col in range(bank_col0, bank_col0 + len(BANK_COLS)):
                    ws.cell(row=b_idx + 2, column=col).fill = fill

    if not extras_df.empty:
        for idx in extras_df.index:
            excel_row = idx + 2
            for col in range(ext_col0, ext_col0 + len(QB_COLS)):
                ws.cell(row=excel_row, column=col).fill = RED

    for col in range(1, ws.max_column + 1):
        try:
            width = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)) + 2
            header = ws.cell(row=1, column=col).value
            if header in ("Split", "Description"):
                width = min(width, 60)  # cap wide text columns
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        except (ValueError, TypeError):
            pass

    ws.freeze_panes = "A2"
    wb.save(output_path)

def gui() -> None:
    class App(ctk.CTk):
        def __init__(self):
            super().__init__()
            self.title("Credit Reconciliation")
            self.minsize(720, 340)
            self.columnconfigure(1, weight=1)
            for i in range(3):
                self.rowconfigure(i, pad=10)
            self._build()

        def _build(self):
            self.bank_var = ctk.StringVar()
            self.qb_var = ctk.StringVar()
            self.out_var = ctk.StringVar()

            ctk.CTkLabel(self, text="Bank credit CSV").grid(row=0, column=0, sticky="w", padx=20)
            ctk.CTkEntry(self, textvariable=self.bank_var).grid(row=0, column=1, sticky="ew", padx=(0, 10))
            ctk.CTkButton(self, text="Browse", command=self._browse_bank).grid(row=0, column=2, padx=10)

            ctk.CTkLabel(self, text="QB credit CSV").grid(row=1, column=0, sticky="w", padx=20)
            ctk.CTkEntry(self, textvariable=self.qb_var).grid(row=1, column=1, sticky="ew", padx=(0, 10))
            ctk.CTkButton(self, text="Browse", command=self._browse_qb).grid(row=1, column=2, padx=10)

            ctk.CTkLabel(self, text="Output Excel").grid(row=2, column=0, sticky="w", padx=20)
            ctk.CTkEntry(self, textvariable=self.out_var).grid(row=2, column=1, sticky="ew", padx=(0, 10))
            ctk.CTkButton(self, text="Save As", command=self._browse_out).grid(row=2, column=2, padx=10)

            run_btn = ctk.CTkButton(self, text="Run", width=120, command=self._run)
            run_btn.grid(row=3, column=0, columnspan=3, pady=30)

        def _browse_bank(self):
            path = fd.askopenfilename(filetypes=[("CSV Files", "*.csv")])
            if path:
                self.bank_var.set(path)

        def _browse_qb(self):
            path = fd.askopenfilename(filetypes=[("CSV Files", "*.csv")])
            if path:
                self.qb_var.set(path)

        def _browse_out(self):
            path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if path:
                self.out_var.set(path)

        def _run(self):
            bank = self.bank_var.get()
            qb = self.qb_var.get()
            out = self.out_var.get()
            if not (bank and qb and out):
                mb.showerror("Missing files", "Please choose all three files.")
                return
            try:
                reconcile(bank, qb, out)
                mb.showinfo("Success", f"Reconciliation complete.\nSaved to {out}")
            except Exception as e:
                mb.showerror("Error", str(e))

    App().mainloop()

if __name__ == "__main__":
    gui()